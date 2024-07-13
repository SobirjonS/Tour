import os
from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated, IsAdminOrReadOnly
from rest_framework.decorators import api_view, authentication_classes, permission_classes

from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from django.core import mail
from django.conf import settings
from django.urls import reverse

import qrcode
import io
from datetime import datetime, date, timedelta
import openpyxl

from . import models
from . import serializers
from . import filters   


# Category
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def create_category(request):
    try:
        name = request.data['name']
        models.Category.objects.create(
            name=name
        )
        return Response("The category has been created", status.HTTP_200_OK)
    except:
        return Response('Bad request', status.HTTP_400_BAD_REQUEST)
    

@api_view(['GET'])
def get_categorys(request):
    categorys = models.Category.objects.all()
    serialiser = serializers.CategorySerializer(categorys, many=True)
    return Response(serialiser.data, status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def update_category(request, pk):
    try:
        category = models.Category.objects.get(pk=pk)
        category.name = request.data['name']
        category.save()
        return Response("The category has been changed", status.HTTP_200_OK)
    except:
        return Response('Did not find such information', status.HTTP_400_BAD_REQUEST)
    

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def delete_category(request, pk):
    try:
        category = models.Category.objects.get(pk=pk)
        category.delete()
        return Response("The category has been deleted", status.HTTP_200_OK)
    except:
        return Response('Did not find such information', status.HTTP_400_BAD_REQUEST)


# Tour
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def create_tour(request):
    user = request.user
    data = request.data
    try:
        images = request.FILES.getlist('image')
        videos = request.FILES.getlist('media')
        services = request.POST.getlist('service')
        category = models.Category.objects.get(id=data['category'])

        tour = models.Tour.objects.create(
            creator=user,
            category=category,
            place_name=data['place_name'],
            title=data['title'],
            body=data['body'],
            cost=data['cost'],
            start_time=data['start_time'],
            end_time=data['end_time'],
            seats=data['seats']
        )

        for image in images:
            models.TourImage.objects.create(
                image=image,
                tour=tour
            )

        for media in videos:
            models.TourMedia.objects.create(
                media=media,
                tour=tour
            )

        for service in services:
            models.TourService.objects.create(
                service=service,
                tour=tour
            )
        return Response('The tour has been created', status.HTTP_200_OK)
    except:
        return Response('Bad request', status.HTTP_400_BAD_REQUEST)


class TourViewSet(ModelViewSet):
    queryset = models.Tour.objects.all()
    serializer_class = serializers.TourSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.TourFilter


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def delete_tour(request, pk):
    try:
        tour = models.Tour.objects.get(pk=pk)
        tour_images = models.TourImage.objects.filter(tour=tour)
        tour_videos = models.TourMedia.objects.filter(tour=tour)

        for image in tour_images:
            image.image.delete(save=False)

        for media in tour_videos: 
            media.media.delete(save=False) 

        tour.delete()
        return Response("The tour has been deleted", status.HTTP_200_OK)
    except:
        return Response('Did not find such information', status.HTTP_400_BAD_REQUEST)
    

@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def update_tour(request, pk):
    try:
        images = request.FILES.getlist('image')
        videos = request.FILES.getlist('media')
        services = request.POST.getlist('service')

        tour = models.Tour.objects.get(pk=pk)

        tour_images = models.TourImage.objects.filter(tour=tour)
        tour_videos = models.TourMedia.objects.filter(tour=tour)
        tour_services = models.TourService.objects.filter(tour=tour)

        serializer = serializers.TourSerializer(tour, data=request.data, partial=(request.method == 'PATCH'))
        if serializer.is_valid():
            serializer.save()

            for image in tour_images:
                image.image.delete(save=False)
                image.delete()

            for image in images:
                models.TourImage.objects.create(
                    image=image,
                    tour=tour
                )

            for media in tour_videos: 
                media.media.delete(save=False)
                media.delete()

            for media in videos:
                models.TourMedia.objects.create(
                    media=media,
                    tour=tour
                )

            for service in tour_services:
                service.delete()

            for service in services:
                models.TourService.objects.create(
                    service=service,
                    tour=tour
                )
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except:
        return Response('Did not find such information', status.HTTP_400_BAD_REQUEST)
    

# Raiting
@api_view(['POST', 'PUT'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def create_or_updare_raiting(request, pk): 
    author = request.user
    tour = models.Tour.objects.get(pk=pk)
    if request.method == "POST":
        try:
            raiting = models.Raiting.objects.get(author=author, tour=tour)
            return Response("Tour has been rated", status.HTTP_400_BAD_REQUEST)
        except models.Raiting.DoesNotExist:
            models.Raiting.objects.create(
                author=author,
                tour=tour,
                grade=request.data['grade']
            )            
            return Response("The raiting has been created", status.HTTP_200_OK)
    if request.method == "PUT":
        try:
            raiting = models.Raiting.objects.get(author=author, tour=tour)
            raiting.grade = request.data['grade']
            raiting.save()
            return Response("The raiting has been updated ", status.HTTP_200_OK)
        except:
            return Response('Did not find such information', status.HTTP_400_BAD_REQUEST)
    

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def delete_raiting(request, pk):  
    try:
        author = request.user
        tour = models.Tour.objects.get(pk=pk)
        raiting = models.Raiting.objects.get(author=author, tour=tour)
        raiting.delete()
        return Response("The raiting has been deleted", status.HTTP_200_OK)
    except:
        return Response('Did not find such information', status.HTTP_400_BAD_REQUEST)
    

# Feedbag
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def create_feedbag(request, pk):
    try:
        author = request.user
        tour = models.Tour.objects.get(pk=pk)
        models.Feedbag.objects.create(
            author=author,
            tour=tour,
            feedbag=request.data['feedbag']
        )            
        return Response("The feedbag has been created", status.HTTP_200_OK)
    except:
        return Response('Did not find such information', status.HTTP_400_BAD_REQUEST)
    

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def update_feedbag(request, pk):
    try:
        feedbag = models.Feedbag.objects.get(pk=pk)   
        if "status" in request.data and request.user.is_admin:
            if request.data['status'] == "1":
                feedbag.status = True
            else:
                feedbag.status = False
        elif 'feedbag' in request.data:
            feedbag.feedbag = request.data['feedbag']
        feedbag.save()
        return Response("The feedbag has been updated", status.HTTP_200_OK)
    except:
        return Response('Did not find such information', status.HTTP_400_BAD_REQUEST)
    

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def delete_feedbag(request, pk):
    try:
        feedbag = models.Feedbag.objects.get(pk=pk)   
        if feedbag.author == request.user or request.user.is_admin:
            feedbag.delete()
            return Response("The feedbag has been deleted", status.HTTP_200_OK)
    except:
        return Response('Did not find such information', status.HTTP_400_BAD_REQUEST)
    

# Booking
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def create_booking(request, pk):
    try:
        buyer = request.user
        tour = models.Tour.objects.get(pk=pk)
        
        requested_seats = int(request.data['seats'])

        if tour.seats >= requested_seats:
            booking = models.Booking.objects.create(
                buyer=buyer,
                tour=tour,
                for_connect=request.data['for_connect'],
                token=request.data['token'],
                seats=requested_seats
            )
            tour.seats = tour.seats - booking.seats
            tour.save()
            return Response("The booking has been created", status.HTTP_200_OK)
        else:
            return Response("Not enough seats available", status.HTTP_400_BAD_REQUEST)
    except models.Tour.DoesNotExist:
        return Response('Tour not found', status.HTTP_404_NOT_FOUND)
    except ValueError:
        return Response('Invalid input for seats', status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response(f'An error occurred: {str(e)}', status.HTTP_400_BAD_REQUEST)
    

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def update_booking(request, pk):
    try:
        user = request.user
        booking = models.Booking.objects.get(pk=pk)  
        tour = booking.tour
        requested_status = int(request.data['status']) 
        if booking.buyer == request.user or request.user.is_admin:
            booking.status = requested_status
            booking.save()
        if requested_status == 1:
            tour.seats = tour.seats - booking.seats
            tour.save()
            r_status = 'Bron qilingan'
        elif requested_status == 2:
            r_status = 'Bron qilingan'
        elif requested_status == 3:
            tour.seats = tour.seats + booking.seats
            tour.save()
            r_status = 'Bron qilingan'

        mail.send_mail(
        'Reset Your Password',
        f'The booking`s status has been updated "{r_status}"',
        settings.EMAIL_HOST_USER,
        [user.email],
        fail_silently=False,
    )
        return Response("The booking has been updated", status.HTTP_200_OK)
    except:
        return Response('Did not find such information', status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def get_booking(request):
    user = request.user
    bookings = models.Booking.objects.filter(buyer=user)
    if bookings:
        serializer = serializers.BookingSerializer(bookings, many=True)
        return Response(serializer.data, status.HTTP_200_OK)
    else:
        return Response('Did not find such information', status.HTTP_400_BAD_REQUEST)
    

class BookingViewSet(ModelViewSet):
    queryset = models.Booking.objects.all()
    serializer_class = serializers.BookingSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.BookingFilter
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminOrReadOnly])
@authentication_classes([TokenAuthentication])
def get_bookings_by_day(request, year, month, day):
    try:
        target_date = date(int(year), int(month), int(day))
        bookings = models.Booking.objects.filter(created_at=target_date)
        serializer = serializers.BookingSerializer(bookings, many=True)
        return Response(serializer.data, status=200)
    except:
        return Response({'error': 'Invalid date format'}, status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminOrReadOnly])
@authentication_classes([TokenAuthentication])
def get_bookings_by_week(request, year, week):
    try:
        start_of_week = datetime.strptime(f'{year}-W{int(week)}-1', "%Y-W%W-%w").date()
        end_of_week = start_of_week + timedelta(days=6)
        
        bookings = models.Booking.objects.filter(created_at__range=[start_of_week, end_of_week])
        serializer = serializers.BookingSerializer(bookings, many=True)
        return Response(serializer.data, status=200)
    except:
        return Response({'error': 'Invalid date format'}, status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminOrReadOnly])
@authentication_classes([TokenAuthentication])
def get_bookings_by_month(request, year, month):
    try:
        bookings = models.Booking.objects.filter(created_at__year=int(year), created_at__month=int(month))
        serializer = serializers.BookingSerializer(bookings, many=True)
        return Response(serializer.data, status=200)
    except:
        return Response({'error': 'Invalid date format'}, status.HTTP_400_BAD_REQUEST)
    

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminOrReadOnly])
@authentication_classes([TokenAuthentication])
def booking_report_by_day(request, year, month, day):
    try:
        target_date = date(int(year), int(month), int(day))
        bookings = models.Booking.objects.filter(created_at=target_date)
        serializer = serializers.BookingSerializer(bookings, many=True)

        if request.method == 'POST':
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Bookings"
            
            headers = ["ID", "For Connect", "Seats", "Buyer", "Tour Title", "Tour Start Time", "Tour End Time", "Token", "Created At", "Status"]
            sheet.append(headers)
            
            for booking in serializer.data:
                tour = booking.get('tour', {})
                row = [
                    booking['id'],
                    booking['for_connect'],
                    booking['seats'],
                    booking['buyer'],
                    tour.get('title', ''),  
                    tour.get('start_time', ''), 
                    tour.get('end_time', ''),
                    booking['token'],
                    booking['created_at'],
                    booking['status'],
                ]
                sheet.append(row)
            
            os.makedirs('reports', exist_ok=True)
            
            filename = f'booking_report_{year}_{month}_{day}-day.xlsx'
            filepath = os.path.join('reports', filename)
            workbook.save(filepath)

            with open(filepath, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={filename}'
                return response
        else:
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(serializer.data)
            qr.make(fit=True)

            img = qr.make_image(fill='black', back_color='white')
            buffer = io.BytesIO()
            img.save(buffer)
            buffer.seek(0)

            return HttpResponse(buffer, content_type='image/png')
    except Exception as e:
        return Response({'error': str(e)}, status.HTTP_400_BAD_REQUEST)
    

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminOrReadOnly])
@authentication_classes([TokenAuthentication])
def bookings_report_by_week(request, year, week):
    try:
        start_of_week = datetime.strptime(f'{year}-W{int(week)}-1', "%Y-W%W-%w").date()
        end_of_week = start_of_week + timedelta(days=6)
        
        bookings = models.Booking.objects.filter(created_at__range=[start_of_week, end_of_week])
        serializer = serializers.BookingSerializer(bookings, many=True)

        if request.method == 'POST':
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Bookings"
            
            headers = ["ID", "For Connect", "Seats", "Buyer", "Tour Title", "Tour Start Time", "Tour End Time", "Token", "Created At", "Status"]
            sheet.append(headers)
            
            for booking in serializer.data:
                tour = booking.get('tour', {})
                row = [
                    booking['id'],
                    booking['for_connect'],
                    booking['seats'],
                    booking['buyer'],
                    tour.get('title', ''),  
                    tour.get('start_time', ''), 
                    tour.get('end_time', ''),
                    booking['token'],
                    booking['created_at'],
                    booking['status'],
                ]
                sheet.append(row)
            
            os.makedirs('reports', exist_ok=True)
            
            filename = f'booking_report_{year}_{week}-week.xlsx'
            filepath = os.path.join('reports', filename)
            workbook.save(filepath)

            with open(filepath, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={filename}'
                return response
        else:
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(serializer.data)
            qr.make(fit=True)

            img = qr.make_image(fill='black', back_color='white')
            buffer = io.BytesIO()
            img.save(buffer)
            buffer.seek(0)

            return HttpResponse(buffer, content_type='image/png')
    except Exception as e:
        return Response({'error': str(e)}, status.HTTP_400_BAD_REQUEST)
    

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminOrReadOnly])
@authentication_classes([TokenAuthentication])
def booking_report_by_month(request, year, month):
    try:
        bookings = models.Booking.objects.filter(created_at__year=int(year), created_at__month=int(month))
        serializer = serializers.BookingSerializer(bookings, many=True)

        if request.method == 'POST':
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Bookings"
            
            headers = ["ID", "For Connect", "Seats", "Buyer", "Tour Title", "Tour Start Time", "Tour End Time", "Token", "Created At", "Status"]
            sheet.append(headers)
            
            for booking in serializer.data:
                tour = booking.get('tour', {})
                row = [
                    booking['id'],
                    booking['for_connect'],
                    booking['seats'],
                    booking['buyer'],
                    tour.get('title', ''),  
                    tour.get('start_time', ''), 
                    tour.get('end_time', ''),
                    booking['token'],
                    booking['created_at'],
                    booking['status'],
                ]
                sheet.append(row)
            
            os.makedirs('reports', exist_ok=True)
            
            filename = f'booking_report_{year}_{month}-month.xlsx'
            filepath = os.path.join('reports', filename)
            workbook.save(filepath)

            with open(filepath, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={filename}'
                return response
        else:
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(serializer.data)
            qr.make(fit=True)

            img = qr.make_image(fill='black', back_color='white')
            buffer = io.BytesIO()
            img.save(buffer)
            buffer.seek(0)

            return HttpResponse(buffer, content_type='image/png')
    except Exception as e:
        return Response({'error': str(e)}, status.HTTP_400_BAD_REQUEST)
    

# Answer
@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminOrReadOnly])
@authentication_classes([TokenAuthentication])
def create_answer(request, pk):
    try:
        user = request.user
        feedbag = models.Feedbag.objects.get(pk=pk)
        answer = request.data['answer']
        models.AnswerFeedbag.objects.create(
            answer=answer,
            feedbag=feedbag,
            author=user
        )
        return Response('The answer has been created', status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status.HTTP_400_BAD_REQUEST)
    

@api_view(['PUT'])
@permission_classes([IsAuthenticated, IsAdminOrReadOnly])
@authentication_classes([TokenAuthentication])
def update_answer(request, pk):
    try:
        user = request.user
        answer = models.AnswerFeedbag.objects.get(pk=pk)
        if user == answer.author:
            answer.answer = request.data['answer']
            answer.save()
            return Response('The answer has been updated', status.HTTP_200_OK)
        return Response('You don`t have access', status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': str(e)}, status.HTTP_400_BAD_REQUEST)
    

@api_view(['DELETE'])
@permission_classes([IsAuthenticated, IsAdminOrReadOnly])
@authentication_classes([TokenAuthentication])
def delete_answer(request, pk):
    try:
        user = request.user
        answer = models.AnswerFeedbag.objects.get(pk=pk)
        if user == answer.author:
            answer.delete()
            return Response('The answer has been deleted', status.HTTP_200_OK)
        return Response('You don`t have access', status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': str(e)}, status.HTTP_400_BAD_REQUEST)
    
 